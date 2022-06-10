"""Module for parsing SCM Community Datasheet Excel files."""

import json
import logging
import uuid
from itertools import chain
from typing import Dict

import requests
from jsonschema.exceptions import ValidationError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from gsy_framework.community_datasheet.exceptions import CommunityDatasheetException
from gsy_framework.community_datasheet.location_converter import (
    LocationConverter, LocationConverterException)
from gsy_framework.community_datasheet.sheet_parsers import (
    CommunityMembersSheetParser, GeneralSettingsSheetParser, LoadSheetParser, ProfileSheetParser,
    PVSheetParser, StorageSheetParser)
from gsy_framework.scenario_validators import scenario_validator

logger = logging.getLogger(__name__)

PROFILE_KEYS_BY_TYPE = {
    "Load": "daily_load_profile",
    "PV": "power_profile",
    "SmartMeter": "smart_meter_profile"
}


class CommunityDatasheetParser:
    """Parser for Community Datasheet files.

    Args:
        filename: the path to open or a file-like object.
    """
    SHEETNAMES = {"General settings", "Community Members", "Load", "PV", "Storage", "Profiles"}

    def __init__(self, filename):
        self._filename = filename

        self.settings: Dict = None
        self.members: Dict = None
        self.loads: Dict = None
        self.pvs: Dict = None
        self.storages: Dict = None
        self.profiles: Dict = None

    def parse(self) -> Dict:
        """Parse the entire datasheet and return its representation as a dictionary."""
        workbook = self._load_workbook(self._filename)

        self._validate_sheetnames(workbook)
        self._parse_sheets(workbook)

        self._fetch_pv_coordinates(self.pvs, self.members)

        self._validate_missing_profiles()

        assets_by_member = self._get_assets_by_member()

        self._merge_profiles_into_assets(self.profiles, assets_by_member)
        grid = self._create_grid(assets_by_member)

        self._validate_grid(grid)

        return {
            "settings": self.settings,
            "grid": grid
        }

    def as_json(self) -> str:
        """Return the JSON representation of the datasheet."""
        return json.dumps(self.parse(), indent=4)

    @staticmethod
    def _load_workbook(filename: str):
        try:
            return load_workbook(filename, read_only=True)
        except InvalidFileException as ex:
            logger.debug("Error while parsing community datasheet: %s", ex)
            raise CommunityDatasheetException(
                "Community Datasheet format not supported. "
                "Please make sure to use a proper Excel .xlsx file.") from ex

    @classmethod
    def _validate_sheetnames(cls, workbook) -> None:
        current_sheetnames = set(workbook.sheetnames)
        if not current_sheetnames == cls.SHEETNAMES:
            raise CommunityDatasheetException(
                f"The datasheet should contain the following sheets: {cls.SHEETNAMES}."
                f"Found: {current_sheetnames}.")

    @staticmethod
    def _validate_grid(grid):
        try:
            scenario_validator(grid)
        except ValidationError as ex:
            raise CommunityDatasheetException(ex) from ex

    def _validate_missing_profiles(self):
        asset_names = []
        # Storage assets do not define profiles so we don't consider them
        asset_items = [assets.values() for assets in [self.loads, self.pvs]]
        for assets in chain.from_iterable(asset_items):
            asset_names.extend(asset["name"] for asset in assets)

        missing_profiles = [
            asset_name for asset_name in asset_names
            if asset_name not in self.profiles
        ]

        if missing_profiles:
            raise CommunityDatasheetException(
                "Each asset must explicitly define an energy profile in the datasheet. "
                f"The following assets do not define a profile: {missing_profiles}.")

    def _parse_sheets(self, workbook) -> None:
        self.settings = GeneralSettingsSheetParser(workbook["General settings"]).parse()
        self.members = CommunityMembersSheetParser(workbook["Community Members"]).parse()
        self.loads = LoadSheetParser(workbook["Load"]).parse()
        self.pvs = PVSheetParser(workbook["PV"]).parse()
        self.storages = StorageSheetParser(workbook["Storage"]).parse()
        self.profiles = ProfileSheetParser(workbook["Profiles"]).parse()

    def _get_assets_by_member(self) -> Dict:
        """Return a mapping between each member and their assets."""
        assets_by_member = {member_name: [] for member_name in self.members}
        asset_items = [assets.items() for assets in [self.loads, self.pvs, self.storages]]
        for member_name, assets in chain.from_iterable(asset_items):
            assets_by_member[member_name].extend(assets)

        if any((missing := member) not in self.members for member in assets_by_member):
            raise CommunityDatasheetException(
                f'Member "{missing}" was defined in one of the asset sheets'  # noqa: F821
                'but not in the "Community Members" sheet.')

        return assets_by_member

    @staticmethod
    def _merge_profiles_into_assets(profiles: Dict, assets_by_member: Dict) -> None:
        """Merge (in-place) each energy profile into the representation of its own asset."""
        for asset in chain.from_iterable(assets_by_member.values()):
            asset_name = asset["name"]
            if asset_name not in profiles:
                continue
            asset_type = asset["type"]
            try:
                profile_key = PROFILE_KEYS_BY_TYPE[asset_type]
            except KeyError as ex:
                raise CommunityDatasheetException(
                    f'Asset type "{asset_type}" is not supported.') from ex
            asset[profile_key] = profiles[asset_name]

    def _create_grid(self, assets_by_member: Dict) -> Dict:
        members_grid = []
        for member_name, assets in assets_by_member.items():
            member_representation = self._create_member_representation(member_name)
            member_representation["children"] = assets
            members_grid.append(member_representation)

        return {
            "name": "Grid Market",
            "tags": None,
            "type": "Area",
            "uuid": str(uuid.uuid4()),
            "children": members_grid
        }

    def _create_member_representation(self, member_name: str) -> Dict:
        return {
            **self.members[member_name],
            **{
                "name": member_name,
                "tags": ["Home"],
                "type": "Area",
                "uuid": str(uuid.uuid4()),
                "children": []
            }}

    @classmethod
    def _fetch_pv_coordinates(cls, pvs_by_member, members_information):
        try:
            location_converter = LocationConverter()
        except LocationConverterException as ex:
            logger.warning(ex)
            return

        with requests.Session() as session:
            for member_name, pv_assets in pvs_by_member.items():
                coordinates = cls._get_member_coordinates(
                    members_information[member_name],
                    location_converter=location_converter,
                    session=session)
                for pv_details in pv_assets:
                    pv_details["geo_tag_location"] = coordinates

    @staticmethod
    def _get_member_coordinates(
            member_details: Dict, location_converter: LocationConverter, session=None):
        address = member_details.get("Location/Address (optional)") or ""
        zip_code = member_details.get("ZIP code") or ""
        full_address = f"{address} {zip_code}"
        if not full_address.strip():
            return None

        try:
            return location_converter.convert(full_address, session=session)
        except LocationConverterException as ex:
            raise CommunityDatasheetException(ex) from ex


def parse_community_datasheet(filename):
    """Parse the content of the community datasheet and return a JSON representation."""
    datasheet = CommunityDatasheetParser(filename)
    output = datasheet.parse()
    print(json.dumps(output, indent=4))


if __name__ == "__main__":
    parse_community_datasheet(
        "/Users/fievelk/Code/work/grid_singularity/Community_DataSheet_v2_ppp.xlsx")

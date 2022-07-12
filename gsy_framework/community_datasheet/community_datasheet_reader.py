"""Module for the reader of the SCM community datasheet."""

import json
import logging
import pathlib
from dataclasses import asdict, dataclass, field
from itertools import chain
from typing import IO, Dict, List, Union

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from gsy_framework.community_datasheet.exceptions import CommunityDatasheetException
from gsy_framework.community_datasheet.sheet_parsers import (
    CommunityMembersSheetParser, GeneralSettingsSheetParser, LoadSheetParser, ProfileSheetParser,
    PVSheetParser, StorageSheetParser)

logger = logging.getLogger(__name__)


@dataclass
class CommunityDatasheet:
    """Dataclass for the parsed contents of the Community Datasheet."""

    settings: Dict
    members: Dict
    loads: Dict[str, List]  # Map member names to their loads
    pvs: Dict[str, List]  # Map member names to their pvs
    storages: Dict[str, List]  # Map member names to their storages
    profiles: Dict
    grid: Dict = field(default_factory=dict)

    def as_json(self) -> str:
        """Return the JSON representation of the datasheet."""
        return json.dumps(asdict(self), indent=4)

    def as_dict(self) -> str:
        """Return the dictionary representation of the datasheet."""
        return asdict(self)

    @property
    def load_assets(self) -> List:
        """Return a list of the load assets defined in the datasheet."""
        return [
            asset
            for member_assets in self.loads.values()
            for asset in member_assets
        ]

    @property
    def pv_assets(self) -> List:
        """Return a list of the PV assets defined in the datasheet."""
        return [
            asset
            for member_assets in self.pvs.values()
            for asset in member_assets
        ]

    @property
    def storage_assets(self) -> List:
        """Return a list of the storage assets defined in the datasheet."""
        return [
            asset
            for member_assets in self.storages.values()
            for asset in member_assets
        ]

    @property
    def all_assets(self) -> List:
        """Return a list of all assets defined in the datasheet."""
        return [
            asset
            for assets_group in (self.load_assets, self.pv_assets, self.storage_assets)
            for asset in assets_group]

    @property
    def assets_by_member(self) -> Dict:
        """Return a mapping between each member and their assets.

        Example:
            {
                "Member 1": [
                    {"name": "Load 1", "type": "Load", "uuid": <some-uuid>},
                    {"name": "PV 1", "type": "PV", "uuid": <some-uuid>}
                ],
                "Member 2": [...]
            }
        """
        assets_by_member = {member_name: [] for member_name in self.members}
        asset_items = [assets.items() for assets in (self.loads, self.pvs, self.storages)]
        for member_name, assets in chain.from_iterable(asset_items):
            assets_by_member[member_name].extend(assets)

        if any((missing := member) not in self.members for member in assets_by_member):
            raise CommunityDatasheetException(
                f'Member "{missing}" was defined in one of the asset sheets'  # noqa: F821
                'but not in the "Community Members" sheet.')

        return assets_by_member


class CommunityDatasheetReader:
    """Reader for Community Datasheet files."""

    SHEETNAMES = {"General settings", "Community Members", "Load", "PV", "Storage", "Profiles"}

    @classmethod
    def read(cls, filename: Union[str, pathlib.Path, IO]) -> Dict:
        """Parse the entire datasheet and return its dataclass.

        Args:
            filename: the path to open or a file-like object.
        """
        workbook = cls._load_workbook(filename)

        cls._validate_sheetnames(workbook)
        datasheet = cls._parse_sheets(workbook)

        return datasheet

    @staticmethod
    def _load_workbook(filename: Union[str, pathlib.Path, IO]):
        try:
            return load_workbook(filename, read_only=True)
        except InvalidFileException as ex:
            logger.debug("Error while parsing community datasheet: %s", ex)
            raise CommunityDatasheetException(
                "Community Datasheet format not supported. "
                "Please make sure to use a proper Excel .xlsx file.") from ex

    @classmethod
    def _validate_sheetnames(cls, workbook) -> None:
        current_sheetnames = set(workbook.sheetnames)
        if not current_sheetnames == cls.SHEETNAMES:
            raise CommunityDatasheetException(
                f"The datasheet should contain the following sheets: {cls.SHEETNAMES}."
                f"Found: {current_sheetnames}.")

    @staticmethod
    def _parse_sheets(workbook) -> CommunityDatasheet:
        return CommunityDatasheet(
            settings=GeneralSettingsSheetParser(workbook["General settings"]).parse(),
            members=CommunityMembersSheetParser(workbook["Community Members"]).parse(),
            loads=LoadSheetParser(workbook["Load"]).parse(),
            pvs=PVSheetParser(workbook["PV"]).parse(),
            storages=StorageSheetParser(workbook["Storage"]).parse(),
            profiles=ProfileSheetParser(workbook["Profiles"]).parse()
        )
